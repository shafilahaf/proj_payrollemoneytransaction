import openpyxl
import collections
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime, timedelta
import logging
import base64
import io
import urllib.parse
import random
from collections import defaultdict
import pytz
from pytz import timezone, all_timezones


_logger = logging.getLogger(__name__)

class PayrollAttendanceLogReportExcelWizard(models.TransientModel):
    _name = "payroll.attendance.log.report.excel.wizard"
    _description = "Payroll Attendance Log Report Excel Wizard"

    department_id = fields.Many2one("payroll.device.department", string="Department", required=True)
    category_id = fields.Many2one("payroll.employee.categories", string="Category")
    employee_id = fields.Many2one("payroll.employees", string="Employee")
    start_date = fields.Date(string='Start Date', required=True)
    end_date = fields.Date(string='End Date', required=True)

    def payroll_attendance_log_report(self):
        """
        Generate Payroll Attendance Log Report Excel
        """
        if not self.start_date or not self.end_date:
            raise ValidationError(_("Please provide both Start Date and End Date"))

        # Create Excel file
        wb = Workbook()
        ws = wb.active

        # Define column widths and row heights
        column_widths = {'A': 9.57, 'B': 2.71, 'C': 0.58, 'D': 2.29, 'E': 3.86, 'F': 11.71,
                        'G': 5.14, 'H': 0.1, 'I': 3.71, 'J': 11.14, 'K': 8.43, 'L': 8.71}
        row_heights = {1: 24, 2: 4.5, 3: 18, 4: 9.75, 5: 18, 6: 8, 7: 18, 8: 7.5}

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

        # Define border styles
        border_style = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )

        # Define cell fills for attendance statuses
        yellow_fill = PatternFill(start_color="dfe62e", end_color="dfe62e", fill_type="solid")  
        fills = {
            'Leave': PatternFill(start_color='5094d9', end_color='5094d9', fill_type='solid'),  # Blue
            'Sick': PatternFill(start_color='d19794', end_color='d19794', fill_type='solid'),   # Pink
            'Permission': PatternFill(start_color='8f7160', end_color='8f7160', fill_type='solid'),  # Brown
            'Day Off': PatternFill(start_color='a5a69f', end_color='a5a69f', fill_type='solid'),  # Grey
            'Device Error': PatternFill(start_color='d69c56', end_color='d69c56', fill_type='solid'),  # Orange
            'Not Taken Leave': PatternFill(start_color='dfe62e', end_color='dfe62e', fill_type='solid'),  # Yellow
            'Medical Reimbursement W/O Off': PatternFill(start_color='dfe62e', end_color='dfe62e', fill_type='solid'),  # Yellow
            'Absence': PatternFill(start_color='dfe62e', end_color='dfe62e', fill_type='solid'),  # Yellow
            'Before Active': PatternFill(start_color='dfe62e', end_color='dfe62e', fill_type='solid'),  # Yellow
            'Shift Default': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),  # Default white
            'Late': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'Forget Checkout': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light Blue
        }

        # Set up title and headers
        ws.title = "ATTENDANCE LOGS"
        ws.sheet_view.showGridLines = False  # Hide gridlines

        # Title
        ws.merge_cells('A1:F2')
        ws['A1'] = "ATTENDANCE LOGS"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

        # Subtitle
        ws.merge_cells('A3:B4')
        ws['A3'] = 'Category'
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
        ws['C3'] = self.category_id.name if self.category_id else ""

        ws.merge_cells('A5:B6')
        ws['A5'] = 'Start Date'
        ws['A5'].alignment = Alignment(horizontal='left', vertical='center')
        ws['C5'] = self.start_date.strftime('%d %B %Y')

        ws.merge_cells('A7:B8')
        ws['A7'] = 'End Date'
        ws['A7'].alignment = Alignment(horizontal='left', vertical='center')
        ws['C7'] = self.end_date.strftime('%d %B %Y')

        # Define table headers
        headers = ["NIK", "Name", "", "", "", "", "", "", "Department", "", "Active Date", ""]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style

        # Merge header cells
        ws.merge_cells('B9:G9')
        ws.merge_cells('I9:J9')
        ws.merge_cells('K9:L9')

        # Add date headers
        current_date = self.start_date
        col_num = len(headers) + 1

        while current_date <= self.end_date:
            col_letter = get_column_letter(col_num)
            # ws[col_letter + '9'] = current_date.strftime('%d/%m/%Y')
            ws[col_letter + '9'] = current_date.strftime('%d %b %Y')
            ws[col_letter + '9'].alignment = Alignment(horizontal='center', vertical='center')
            ws[col_letter + '9'].font = Font(bold=True)
            ws[col_letter + '9'].border = border_style
            ws.column_dimensions[col_letter].width = 18.29
            col_num += 1
            current_date += timedelta(days=1)

        # Define domain filters
        domain = [('date', '>=', self.start_date), ('date', '<=', self.end_date)]

        if self.department_id:
            domain.append(('employee_id.department_id', '=', self.department_id.id))
        if self.category_id:
            domain.append(('employee_id.category_id', '=', self.category_id.id))
        if self.employee_id:
            domain.append(('employee_id', '=', self.employee_id.id))

        # Fetch attendance logs and employees
        attendance_logs = self.env['payroll.attendance.log'].search(domain, order='employee_id')
        attendance_logs = sorted(attendance_logs, key=lambda log: log.employee_id.name)
        
        # Initialize the employee_data only for employees with attendance logs
        employee_data = defaultdict(lambda: {"attendance": {}, "info": {}})

        # Define the relevant time zones
        local_timezone = pytz.timezone('Asia/Jakarta')  # GMT+7
        utc_timezone = pytz.utc

        for log in attendance_logs:
            employee = log.employee_id
            date_str = log.date.strftime('%d/%m/%Y')

            if employee not in employee_data:
                # Add employee info to employee_data
                employee_data[employee]["info"] = {
                    "nik": employee.nik or "N/A",
                    "name": employee.name or "N/A",
                    "department": employee.department_id.name if employee.department_id else "N/A",
                    "active_date": employee.active_date.strftime('%d %B %Y') if employee.active_date else "N/A"
                }


            # Check if the attendance log date is before the employee's active date
            # if log.date < employee.active_date:
            #     # Wanna check if starttime not 00:00
            #     start_time_utc = log.start_time.replace(tzinfo=utc_timezone)
            #     start_time_local = start_time_utc.astimezone(local_timezone)
            #     start_time_not_00 = start_time_local.strftime('%H:%M') != '00:00'
            #     if log.start_time and start_time_not_00:
            #         # Convert `start_time` from UTC to local timezone (GMT+7)
            #         start_time_utc = log.start_time.replace(tzinfo=utc_timezone)
            #         start_time_local = start_time_utc.astimezone(local_timezone)
            #         shift_time = start_time_local.strftime('%H:%M')
            #         fill_color = fills["Before Active"]  # Highlight in yellow
            #     else:
            #         shift_time = "Before Active "
            #         fill_color = fills["Before Active"]  
            # elif log.time_off_type:
            #     # Map time_off_type to a readable label
            #     shift_time = {
            #         '1': 'Leave',
            #         '2': 'Sick',
            #         '3': 'Permission',
            #         '4': 'Day Off',
            #         '5': 'Device Error',
            #         '6': 'Not Taken Leave',
            #         '7': 'Medical Reimbursement W/O Off',
            #         '99': 'Absence'
            #     }.get(log.time_off_type, 'Unknown')
            # elif log.start_time:
            #     # Convert `start_time` from UTC to local timezone (GMT+7)
            #     start_time_utc = log.start_time.replace(tzinfo=utc_timezone)
            #     start_time_local = start_time_utc.astimezone(local_timezone)

            #     # Ensure that the time is formatted correctly as 'HH:MM'
            #     shift_time = start_time_local.strftime('%H:%M')  # Use '%H:%M:%S' for full time format
            # else:
            #     shift_time = "Forget Checkout"

            # NEW
            if log.end_time == False:
                shift_time = "Forget Checkout"
            elif log.time_off_type == False:
                start_time_utc = log.start_time.replace(tzinfo=utc_timezone)
                start_time_local = start_time_utc.astimezone(local_timezone)
                shift_time = start_time_local.strftime('%H:%M')
            else:
                shift_time = {
                    '1': 'Leave',
                    '2': 'Sick',
                    '3': 'Permission',
                    '4': 'Day Off',
                    '5': 'Device Error',
                    '6': 'Not Taken Leave',
                    '7': 'Medical Reimbursement W/O Off',
                    '99': 'Absence'
                }.get(log.time_off_type, 'Unknown')
            # NEW

            # Apply the corresponding fill color based on the shift_time/status
            if log.is_latelog or log.is_late_2 or log.is_late_3 or log.is_late_4:
                fill_color = fills.get(shift_time, fills["Late"])
            else:
                fill_color = fills.get(shift_time, fills["Shift Default"])

            # Store the attendance information in the employee_data dictionary
            employee_data[employee]["attendance"][date_str] = {
                "status": shift_time,
                "fill": fill_color
            }

        # Populate data rows
        row_num = 10
        for employee, data in employee_data.items():
            ws.cell(row=row_num, column=1, value=data["info"].get("nik", "N/A")).border = border_style
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=7)
            ws.merge_cells(start_row=row_num, start_column=9, end_row=row_num, end_column=10)
            ws.merge_cells(start_row=row_num, start_column=11, end_row=row_num, end_column=12)

            ws.cell(row=row_num, column=2, value=data["info"].get("name", "N/A")).border = border_style
            ws.cell(row=row_num, column=9, value=data["info"].get("department", "N/A")).border = border_style
            ws.cell(row=row_num, column=11, value=data["info"].get("active_date", "N/A")).border = border_style

            for col in range(2, 8):
                ws.cell(row=row_num, column=col).border = border_style
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(9, 11):
                ws.cell(row=row_num, column=col).border = border_style
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center', vertical='center')

            for col in range(11, 13):
                ws.cell(row=row_num, column=col).border = border_style
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center', vertical='center')

            # Populate attendance data
            col_num = len(headers) + 1
            current_date = self.start_date

            while col_num <= ws.max_column:
                col_letter = get_column_letter(col_num)
                date_str = current_date.strftime('%d/%m/%Y')    
                attendance_data = data["attendance"].get(date_str, {"status": '', "fill": fills["Absence"]})
                attendance_status = attendance_data["status"]
                fill_color = attendance_data["fill"]

                cell = ws.cell(row=row_num, column=col_num, value=attendance_status)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style
                cell.fill = fill_color


                if current_date < employee.active_date:
                    if attendance_status and attendance_status != "Before Active":
                        cell.value = attendance_status
                        cell.fill = fills.get(attendance_status, yellow_fill)  # Use the status fill color or yellow if it's before active date
                    else:
                        cell.value = "Before Active"
                        cell.fill = yellow_fill
                elif attendance_status == '':
                    cell.value = "Absence"
                    cell.fill = fills["Absence"]
                # else:
                #     cell.fill = fills.get(attendance_status, fills["Shift Default"])
     
                current_date += timedelta(days=1)
                col_num += 1
            row_num += 1


        # Add a row below the last data row for the note
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)  # Merge columns A to E for the note
        ws.cell(row=row_num, column=1, value="Note:").font = Font(bold=True)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='left', vertical='center')

        row_num += 2  # Move to the next row for the grid

        # Define colors for each label
        colors = {
            "LATE": "FF0000",  # Red
            "SICK": "d19794",  # Purple
            "DAY OFF": "a5a69f",  # Grey
            "LEAVE": "5094d9",  # Cyan
            "PERMISSION": "8f7160",  # Light brown
            "DEVICE ERROR": "d69c56",  # Brown
            "ABSENCE": "dfe62e",  # Yellow
            "BEFORE ACTIVE": "dfe62e",  # Yellow (same as ABSENCE)
            "FORGET CHECKOUT": "ADD8E6"  # Light blue
        }

        # Labels in the required grid format
        grid_labels = [
            ["LATE", "LEAVE", "ABSENCE"],
            ["SICK", "PERMISSION", "BEFORE ACTIVE"],
            ["DAY OFF", "DEVICE ERROR", "FORGET CHECKOUT"]
        ]

        column_spacing = 2
        row_spacing = 2

        # Mapping grid labels to their positions
        grid_mapping = [
            {"label": "LATE", "start_col": "A", "start_merge": None},
            {"label": "LEAVE", "start_col": "D", "start_merge": "F"},
            {"label": "ABSENCE", "start_col": "I", "start_merge": "K"},

            {"label": "SICK", "start_col": "A", "start_merge": None},
            {"label": "PERMISSION", "start_col": "D", "start_merge": "F"},
            {"label": "BEFORE ACTIVE", "start_col": "I", "start_merge": "K"},

            {"label": "DAY OFF", "start_col": "A", "start_merge": None},
            {"label": "DEVICE ERROR", "start_col": "D", "start_merge": "F"},
            {"label": "FORGET CHECKOUT", "start_col": "I", "start_merge": "K"}
        ]

        # Iterate over the grid_mapping
        for idx, item in enumerate(grid_mapping):
            cell_value = item["label"]
            cell_color = colors[cell_value]
            
            # Determine row position based on the iteration and row spacing
            row_start = row_num + (idx // 3 * (1 + row_spacing))  # Every three items, move to the next row

            # Get the column based on the label position
            start_column_letter = item["start_col"]
            start_column = ws[start_column_letter + str(row_start)].column

            # If merging cells is needed
            if item["start_merge"]:
                end_column_letter = item["start_merge"]
                end_column = ws[end_column_letter + str(row_start)].column

                # Merge cells across the defined columns
                ws.merge_cells(start_row=row_start, start_column=start_column, end_row=row_start, end_column=end_column)

                for col in range(start_column, end_column + 1):
                    cell = ws.cell(row=row_start, column=col)
                    cell.border = border_style  # Apply border to every cell in the merged range
            else:
                # Apply border to the single cell
                cell = ws.cell(row=row_start, column=start_column)
                cell.border = border_style

            # Set the value and format of the cell
            cell = ws.cell(row=row_start, column=start_column, value=cell_value)
            
            # Apply font, alignment, color, and border to the cell
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
            cell.border = border_style

        # Move the row_num pointer for future usage if needed
        row_num = row_start + 3 + row_spacing

        # Save Excel file to binary field
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        # Encode data to base64
        data_base64 = base64.b64encode(data)

        name_file = "ATTENDANCE LOG REPORT.xlsx"

        # Save Excel file to attachment
        attachment = self.env['ir.attachment'].create({
            'name': name_file,
            'type': 'binary',
            'datas': data_base64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return attachment
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s/%s' % (attachment.id, urllib.parse.quote(attachment.name)),
            'target': 'self'
        }
    
    def payroll_PH_internal_attendance_log_report(self):
        """
        Generate Payroll PH Internal Attendance Log Report Excel
        """
        if not self.start_date or not self.end_date:
            raise ValidationError(_("Please provide both Start Date and End Date"))
        
        # Get the start_date and end_date from the Odoo model
        start_date = self.start_date
        end_date = self.end_date

        # Create Excel file
        wb = Workbook()
        ws = wb.active

        # Set up dynamic column widths based on content length
        def set_auto_width(sheet, min_width=10):
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                max_length = max(max_length, min_width)  # Ensure a minimum width
                column_letter = get_column_letter(column_cells[0].column)
                sheet.column_dimensions[column_letter].width = max_length

        # Define border styles
        border_style = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )

        # Row 2: Date range in the first cell
        date_range = f"{start_date.strftime('%d %B %Y')} - {end_date.strftime('%d %B %Y')}"
        ws['A2'].value = date_range
        ws['A2'].font = Font(bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 3: Static header "Name" and "Active Date"
        ws['A3'].value = "NAME"
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['B3'].value = "ACTIVE DATE"
        ws['B3'].font = Font(bold=True)
        ws['B3'].alignment = Alignment(horizontal='center', vertical='center')

        # Generate dynamic date headers based on start_date and end_date
        current_date = start_date
        col_idx = 3  # Start from column C (which is index 3) as B is now for "Active Date"
        while current_date <= end_date:
            day_name = current_date.strftime('%A')
            day_number = current_date.strftime('%d')
            
            # Set day name
            ws.cell(row=2, column=col_idx).value = day_name
            ws.cell(row=2, column=col_idx).font = Font(bold=True)
            ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            
            # Set day number in the next row
            ws.cell(row=3, column=col_idx).value = day_number
            ws.cell(row=3, column=col_idx).font = Font(bold=True)
            ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            
            current_date += timedelta(days=1)
            col_idx += 1

        # Row 2 and 3: Additional headers after the date columns
        additional_headers = ['Day Off', 'Sick', 'Permission', 'Absence', 'Leave', 'ND', 'Late', 
                            'SH Hours', 'LH Hours', 'Regular OT', 'SH OT', 'LH OT']
        for header in additional_headers:
            ws.cell(row=2, column=col_idx).value = header
            ws.cell(row=2, column=col_idx).font = Font(bold=True)
            ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            
            # Merge cells for additional headers
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
            
            col_idx += 1
        
        # Apply borders
        for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=col_idx - 1):
            for cell in row:
                cell.border = border_style
        
        domain = []
        if self.department_id:
            domain.append(('employee_id.department_id', '=', self.department_id.id))
        if self.category_id:
            domain.append(('employee_id.category_id', '=', self.category_id.id))
        if self.employee_id:
            domain.append(('employee_id', '=', self.employee_id.id))
        if self.start_date:
            domain.append(('date', '>=', self.start_date))
        if self.end_date:
            domain.append(('date', '<=', self.end_date))
        
        attendance_log = self.env['payroll.attendance.log'].search(domain)
        grouped_logs = defaultdict(list)
        for log in attendance_log:
            grouped_logs[log.employee_id].append(log)
        
        data_row_idx = 4
        for employee, logs in grouped_logs.items():
            # Set employee name
            ws.cell(row=data_row_idx, column=1).value = employee.name
            ws.cell(row=data_row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Set active date
            ws.cell(row=data_row_idx, column=2).value = employee.active_date
            ws.cell(row=data_row_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')

            # Set shifts in corresponding columns
            date_to_shift = {}
            for log in logs:
                if not log.end_time:
                    shift_time = 'Forget Checkout'
                elif not log.time_off_type:
                    # shift_timezone_str = log.shift_id.timezone  
                    # shift_timezone = pytz.timezone(shift_timezone_str if shift_timezone_str else 'UTC')
                    gmt_plus_8_offset = timedelta(hours=8)
                    utc_plus_7_offset = timedelta(hours=7)
                    if log.shift_id.timezone == 'UTC+7':
                        utc_plus_7_timezones = [tz for tz in pytz.all_timezones if pytz.timezone(tz).utcoffset(None) == utc_plus_7_offset]
                        random_timezone = random.choice(utc_plus_7_timezones)
                    elif log.shift_id.timezone == 'GMT' or log.shift_id.timezone == 'GMT+8':
                        gmt_or_gmt_plus_8_timezones = [
                            tz for tz in pytz.all_timezones 
                            if tz.startswith('GMT') or pytz.timezone(tz).utcoffset(None) == gmt_plus_8_offset
                        ]
                        random_timezone = random.choice(gmt_or_gmt_plus_8_timezones)
                    else:
                        random_timezone = 'UTC'
                    
                    shift_timezone_str = random_timezone
                    shift_timezone = pytz.timezone(shift_timezone_str)
                    
                    start_time_local = log.start_time.astimezone(shift_timezone)
                    end_time_local = log.end_time.astimezone(shift_timezone) if log.end_time else None
                    
                    shift_utc_offset = start_time_local.utcoffset().total_seconds() / 3600
                    
                    start_time_adjusted = start_time_local + timedelta(hours=shift_utc_offset)
                    end_time_adjusted = end_time_local + timedelta(hours=shift_utc_offset) if end_time_local else None
                    
                    start_time_str = start_time_adjusted.strftime("%H:%M")
                    end_time_str = end_time_adjusted.strftime("%H:%M") if end_time_adjusted else ""
                    
                    shift_time = f"{start_time_str} - {end_time_str}"
                     
                    if log.shift_id.color2:
                        hex_color = log.shift_id.color2.lstrip('#')  # Remove the '#' if it exists
                        fill_shift = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                    else:
                        fill_shift = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                else:
                    shift_time = dict(log._fields['time_off_type'].selection).get(log.time_off_type, '')
                
                if log.date in date_to_shift:
                    date_to_shift[log.date].append(shift_time)
                else:
                    date_to_shift[log.date] = [shift_time]
                
            # Combine shift times for each date and assign to the cell
            for date, shifts in date_to_shift.items():
                date_to_shift[date] = ", ".join(shifts)
                
            # Define the fill styles for each case
            fill_late = PatternFill(start_color='e06760', end_color='e06760', fill_type='solid')  # Red 
            fill_leave = PatternFill(start_color='5094d9', end_color='5094d9', fill_type='solid')  # Blue
            fill_absence = PatternFill(start_color='dfe62e', end_color='dfe62e', fill_type='solid')  # Yellow 
            fill_before_active = PatternFill(start_color='dfe62e', end_color='dfe62e', fill_type='solid')  # Yellow 
            fill_day_off = PatternFill(start_color='a5a69f', end_color='a5a69f', fill_type='solid')  # Grey 
            fill_device_error = PatternFill(start_color='d69c56', end_color='d69c56', fill_type='solid')  # Orange 
            fill_forget_checkout = PatternFill(start_color='1be1f7', end_color='1be1f7', fill_type='solid')  # Light Blue
            fill_permission = PatternFill(start_color='8f7160', end_color='8f7160', fill_type='solid')  # Brown
            fill_sick = PatternFill(start_color='d19794', end_color='d19794', fill_type='solid')  # Pink
                        
            # Fill in the shift data for the corresponding date columns
            current_date = start_date
            col_idx = 3  # Start from column C
            while current_date <= end_date:
                cell = ws.cell(row=data_row_idx, column=col_idx)

                if current_date in date_to_shift:
                    value = "".join(date_to_shift[current_date])
                    cell.value = value

                    # Apply fill based on specific conditions
                    if "Forget Checkout" in value:
                        cell.fill = fill_forget_checkout
                    elif "Device Error" in value:
                        cell.fill = fill_device_error
                    elif "Late" in value:
                        cell.fill = fill_late
                    elif "Leave" in value:
                        cell.fill = fill_leave
                    elif "Absence" in value:
                        cell.fill = fill_absence
                    elif "Before Active" in value:
                        cell.fill = fill_before_active
                    elif "Day Off" in value:
                        cell.fill = fill_day_off
                    elif "Permission" in value:
                        cell.fill = fill_permission
                    elif "Sick" in value:
                        cell.fill = fill_sick
                    else:
                        cell.fill = fill_shift
                        
                elif current_date < employee.active_date:
                    cell.value = "Before Active"
                    cell.fill = fill_before_active
                else:
                    cell.value = "Absence"
                    cell.fill = fill_absence

                cell.alignment = Alignment(horizontal='center', vertical='center')
                current_date += timedelta(days=1)
                col_idx += 1
                
                # Adding value another column
                total_late = False
                late_mins = 0
                night_diff = 0
                holiday_sh = 0
                sh_ot = 0
                holiday_lh = 0
                lh_ot = 0
                absence = 0
                day_off = 0
                leave = 0
                permission = 0
                sick = 0
                
                # Calculate late
                if any([log.is_latelog for log in logs]) or any([log.is_late_2 for log in logs]) or \
                    any([log.is_late_3 for log in logs]) or any([log.is_late_4 for log in logs]):
                        total_late = True
                        
                # Calculate late minutes
                for log in logs:
                    if log.late_minutes:
                        late_mins += log.late_minutes
                        
                # Calculate night differential (ND)
                for log in logs:
                    if log.is_night_diff and log.time_off_type == False:
                        night_diff = 8
                        
                # Calculate holidays
                start_day = start_date.day
                start_month = start_date.month
                end_day = end_date.day
                end_month = end_date.month
                holiday_all = self.env['payroll.holiday'].search([])  # Fetch all holidays
                holiday_filter = holiday_all.filtered(lambda h: (
                    (h.date.month == start_month and h.date.day >= start_day) or
                    (h.date.month > start_month and h.date.month < end_month) or
                    (h.date.month == end_month and h.date.day <= end_day)
                ))
                holidays = holiday_filter.sorted(key=lambda h: h.date)[0] if holiday_filter else None

                for holiday in holidays:
                    if holiday.date in date_to_shift:
                        if holiday.type == "2":
                            holiday_sh = 8
                            sh_ot = 3
                        elif holiday.type == "1":
                            holiday_lh = 8
                            lh_ot = 3
                            
                # Calculate absence, day off, leave, permission, sick
                if not employee.category_id.without_attendance_logs:
                    lattlog = self.env['payroll.attendance.log'].search([
                        ('employee_id', '=', employee.id),
                        ('date', '>=', start_date),
                        ('date', '<=', end_date)
                    ])
                    ltotdays = (end_date - start_date).days + 1
                    if lattlog:
                        tot_absence = ltotdays - len(lattlog)
                        absence = tot_absence
                        day_off = len([log for log in lattlog if log.time_off_type == '4'])  # Day Off
                        leave = len([log for log in lattlog if log.time_off_type == '1'])   # Leave
                        permission = len([log for log in lattlog if log.time_off_type == '3'])  # Permission
                        sick = len([log for log in lattlog if log.time_off_type == '2'])   # Sick
                    else:
                        tor = self.env['payroll.time.off.request'].search([
                            ('created_by', '=', employee.user_id),
                            ('start_date', '>=', start_date),
                            ('end_date', '<=', end_date),
                            ('request_type', '=', '99'),
                            ('status', '=', '5')
                        ])
                        ltotdays = (end_date - start_date).days + 1
                        absence = ltotdays - len(tor)
                        day_off = 0
                        leave = 0
                        permission = 0
                        sick = 0
                # Calculate Regular OT
                total_masuk = len(logs) - absence - day_off - leave - permission - sick
                reg_ot = total_masuk * 3 if total_masuk > 9 else 0

                        
                ws.cell(row=data_row_idx, column=col_idx).value = day_off
                ws.cell(row=data_row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 1).value = sick
                ws.cell(row=data_row_idx, column=col_idx + 1).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 2).value = permission
                ws.cell(row=data_row_idx, column=col_idx + 2).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 3).value = absence
                ws.cell(row=data_row_idx, column=col_idx + 3).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 4).value = leave
                ws.cell(row=data_row_idx, column=col_idx + 4).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 5).value = night_diff
                ws.cell(row=data_row_idx, column=col_idx + 5).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 6).value = late_mins
                ws.cell(row=data_row_idx, column=col_idx + 6).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 7).value = holiday_sh
                ws.cell(row=data_row_idx, column=col_idx + 7).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 8).value = holiday_lh
                ws.cell(row=data_row_idx, column=col_idx + 8).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 9).value = reg_ot
                ws.cell(row=data_row_idx, column=col_idx + 9).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 10).value = sh_ot
                ws.cell(row=data_row_idx, column=col_idx + 10).alignment = Alignment(horizontal='center', vertical='center')
                
                ws.cell(row=data_row_idx, column=col_idx + 11).value = lh_ot
                ws.cell(row=data_row_idx, column=col_idx + 11).alignment = Alignment(horizontal='center', vertical='center')

            # Increment row index for the next employee
            data_row_idx += 1
    
        # Dictionary
            shiftname_color = {}
            for log in attendance_log:
                if not log.time_off_type:
                    if log.shift_id:
                        shift_name = log.shift_id.name
                        shift_color = log.shift_id.color2.lstrip('#') if log.shift_id.color2 else '000000'
                        if shift_name not in shiftname_color:
                            shiftname_color[shift_name] = shift_color

        # Output
        shift_row_idx = data_row_idx + 8
        ws.cell(row=shift_row_idx, column=1).value = "Shift List:"
        ws.cell(row=shift_row_idx, column=1).font = Font(bold=True)
        ws.cell(row=shift_row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')

        for i, (shift_name, color2) in enumerate(shiftname_color.items()):
            cell = ws.cell(row=shift_row_idx + i + 1, column=1)
            cell.value = shift_name
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
            
        note_row_idx = data_row_idx + 2
        ws.cell(row=note_row_idx, column=1).value = "Note:"
        ws.cell(row=note_row_idx, column=1).font = Font(bold=True)
        ws.cell(row=note_row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')


        notes = [
            ("Sick", "d19794", "Permission", "8f7160"),  # Pink | Brown
            ("Late", "e06760", "Leave", "5094d9"),  # Red | Blue
            ("Absence and Before Active", "dfe62e", "Day off", "a5a69f"),  # Yellow | Grey
            ("Device Error", "d69c56", "Forget Checkout", "1be1f7")  # Orange | Light Blue
        ]

        for i, (text1, color1, text2, color2) in enumerate(notes):
            # First note
            cell1 = ws.cell(row=note_row_idx + i + 1, column=1)
            cell1.value = text1
            cell1.alignment = Alignment(horizontal='left', vertical='center')
            cell1.fill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')

            # Second note
            cell2 = ws.cell(row=note_row_idx + i + 1, column=2)
            cell2.value = text2
            cell2.alignment = Alignment(horizontal='left', vertical='center')
            cell2.fill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
        

        # Set autofit widths after filling in the values
        set_auto_width(ws)
        
        # Save Excel file to binary field
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        # Encode data to base64
        data_base64 = base64.b64encode(data)

        name_file = "PH INTERNAL ATTENDANCE LOG REPORT.xlsx"

        # Save Excel file to attachment
        attachment = self.env['ir.attachment'].create({
            'name': name_file,
            'type': 'binary',
            'datas': data_base64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return attachment
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s/%s' % (attachment.id, urllib.parse.quote(attachment.name)),
            'target': 'self'
        }

    
    def payroll_PH_report_new_attendance_log_report(self):
        """
        Generate Payroll PH Report New Attendance Log Report Excel
        """
        if not self.start_date or not self.end_date:
            raise ValidationError(_("Please provide both Start Date and End Date"))
        
         # Get the start_date and end_date from the Odoo model
        start_date = self.start_date
        end_date = self.end_date

        def convert_time_based_on_timezone(time, tz_name):
            # Define timezone mappings
            tz_mappings = {
                'UTC+7': 'Asia/Manila',
                'GMT+8': 'Asia/Singapore',
                'GMT': 'Europe/London'
            }

            # Default to UTC if timezone not found
            tz_name = tz_mappings.get(tz_name, 'UTC')
            
            # Convert naive datetime to aware datetime in UTC
            utc_time = time.replace(tzinfo=timezone('UTC'))
            
            # Convert to the specified timezone
            local_time = utc_time.astimezone(timezone(tz_name))
            return local_time
        
        domain = []
        if self.department_id:
            domain.append(('employee_id.department_id', '=', self.department_id.id))
        if self.category_id:
            domain.append(('employee_id.category_id', '=', self.category_id.id))
        if self.employee_id:
            domain.append(('employee_id', '=', self.employee_id.id))
        if self.start_date:
            domain.append(('date', '>=', self.start_date))
        if self.end_date:
            domain.append(('date', '<=', self.end_date))
        
        attendance_log = self.env['payroll.attendance.log'].search(domain)
        
        # Example employee names
        employee_names = list(set(log.employee_id.name for log in attendance_log if log.employee_id))

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Define styles
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # Set column widths
        ws.column_dimensions['A'].width = 16.57
        for col in range(2, 12):  # B to K
            ws.column_dimensions[get_column_letter(col)].width = 13

         # Set auto width for columns L to AH after header fitting
        def auto_fit_column_width(ws, min_col, max_col):
            for col in range(min_col, max_col + 1):
                max_length = 0
                col_letter = get_column_letter(col)
                for row in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[col_letter].width = max_length + 2  # Adjust width for padding
    
        # Header titles
        employee_headers = ["Date", "Day", "Shift", "IN", "OUT", "IN", "OUT", "HOUR", "SCHEDULE", "REMARKS", "LATE"]
        summary_headers = ["Regular WD", "VL", "SL", "Late/UT (Mins)", "Absent (days)", "Reg Hrs OT", "RD Hrs.", "RDTO",
                        "ND", "Special Holiday", "OT Special Holiday", "RD/Special Holiday", "Regular Holiday",
                        "OT/Regular Holiday", "RD/Special Holiday", "Adjustment", "Regular WD (blue color)", "VL", "SL",
                        "Late/UT (Mins)", "Absent (days)", "Reg Hrs OT", "ND"]
        
        def calculate_night_diff(log):
            if log.is_night_diff and log.time_off_type:
                shift_start_hour = int(log.shift_id.start_time)
                if shift_start_hour == 19:
                    return 6
                elif shift_start_hour == 21:
                    return 7
                elif shift_start_hour == 22:
                    return 8
                else:
                    return 0
            return 0
        
        summary_values = [
            sum(1 if log.time_off_type == False else 0 for log in attendance_log) * 4, #Reguler WD
            sum(1 if log.time_off_type in ("1", "3") else 0 for log in attendance_log), #Total VL
            sum(1 if log.time_off_type == "2" else 0 for log in attendance_log),#TotalSL
            sum(log.late_minutes if log.late_minutes != False else 0 for log in attendance_log),#LateMins
            sum(1 if log.time_off_type == "99" else 0 for log in attendance_log), #Absence
            "",
            "",
            "",
            sum(calculate_night_diff(log) for log in attendance_log),#NightDiff
            sum(8 if log.holiday_type =="2" else 0 for log in attendance_log), #Special Holiday
            "",
            "",
            sum(8 if log.holiday_type =="1" else 0 for log in attendance_log), #Reguler Holiday
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]   
        
        # Function to generate date range
        def generate_date_range(start_date, end_date):
            date_list = []
            current_date = start_date
            while current_date <= end_date:
                date_list.append(current_date)
                current_date += timedelta(days=1)
            return date_list

         # Function to fill in employee data block
        def fill_employee_block(start_row, date_range, employee_name, summary_values):
            # Merged blank row
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=11)

            # Employee name
            ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=11)
            ws.cell(row=start_row + 1, column=1).value = employee_name
            ws.cell(row=start_row + 1, column=1).alignment = center_align
            ws.cell(row=start_row + 1, column=1).font = bold_font

            # Fill in the employee headers
            for col_num, header in enumerate(employee_headers, 1):
                ws.cell(row=start_row + 2, column=col_num).value = header
                ws.cell(row=start_row + 2, column=col_num).alignment = center_align
                ws.cell(row=start_row + 2, column=col_num).font = bold_font
                ws.cell(row=start_row + 2, column=col_num).border = thin_border

            # Fill in the attendance data based on the date range
            current_row = start_row + 3
            for date in date_range:
                matching_logs = attendance_log.filtered(lambda log: log.date == date and log.employee_id.name == employee_name)
                # matching_log = attendance_log.filtered(lambda log: log.date == date)
                
                if matching_logs:
                    log = matching_logs[0]  # Assuming one log entry per date

                    ws.cell(row=current_row, column=1).value = date.strftime("%d/%b/%Y")  # Format date
                    ws.cell(row=current_row, column=2).value = date.strftime("%a")  # Day of the week

                    shift = log.shift_id
                    ws.cell(row=current_row, column=3).value = shift.name if shift else ''
                    ws.cell(row=current_row, column=3).fill = PatternFill(start_color=shift.color2.lstrip('#') if shift and shift.color2 else "FFFFFF", end_color=shift.color2.lstrip('#') if shift and shift.color2 else "FFFFFF", fill_type="solid")

                    # Offsets for timezones
                    if log.start_time and log.end_time and shift:
                        local_start_time = convert_time_based_on_timezone(log.start_time, shift.timezone)
                        local_end_time = convert_time_based_on_timezone(log.end_time, shift.timezone)
                        ws.cell(row=current_row, column=4).value = local_start_time.strftime("%I:%M %p")
                        ws.cell(row=current_row, column=7).value = local_end_time.strftime("%I:%M %p")
                    else:
                        ws.cell(row=current_row, column=4).value = ''
                        ws.cell(row=current_row, column=7).value = ''

                    ws.cell(row=current_row, column=8).value = "8.00" if log else "0.00"
                    time_off_label = log._get_time_off_type_label() if log else ''
                    ws.cell(row=current_row, column=10).value = time_off_label
                    ws.cell(row=current_row, column=11).value = log.late_minutes if log.late_minutes else ''

                    for col_num in range(1, 12):
                        ws.cell(row=current_row, column=col_num).border = thin_border
                else:
                    # Handle case where no logs are found for this date and employee
                    ws.cell(row=current_row, column=1).value = date.strftime("%d/%b/%Y")
                    ws.cell(row=current_row, column=2).value = date.strftime("%a")
                    ws.cell(row=current_row, column=3).value = ''
                    ws.cell(row=current_row, column=4).value = ''
                    ws.cell(row=current_row, column=7).value = ''
                    ws.cell(row=current_row, column=8).value = "0.00"
                    ws.cell(row=current_row, column=10).value = ''
                    ws.cell(row=current_row, column=11).value = ''
                    for col_num in range(1, 12):
                        ws.cell(row=current_row, column=col_num).border = thin_border

                current_row += 1
                
                # ws.cell(row=current_row, column=1).value = date.strftime("%d/%b/%Y")  # Format date
                # ws.cell(row=current_row, column=2).value = date.strftime("%a")  # Day of the week
                
                # matching_shift = attendance_log.filtered(lambda log: log.date == date).shift_id
                # shift_name = matching_shift.name if len(matching_shift) == 1 else ''
                # ws.cell(row=current_row, column=3).value = shift_name
                
                # matching_shift_color = attendance_log.filtered(lambda log: log.date == date).shift_id
                # hex_color = matching_shift_color.color2.lstrip('#') if matching_shift_color.color2 else "FFFFFF"
                # shift_color = hex_color if len(matching_shift_color) == 1 else "FFFFFF"
                # ws.cell(row=current_row, column=3).fill = PatternFill(start_color=shift_color, end_color=shift_color, fill_type="solid")
            
                # # Offsets for timezones
                # gmt_plus_8_offset = timedelta(hours=8)
                # utc_plus_7_offset = timedelta(hours=7)
                # matching_starttime = attendance_log.filtered(lambda log: log.date == date).start_time
                # matching_starttime2 = attendance_log.filtered(lambda log: log.date == date).end_time
                # if matching_starttime and len(matching_shift) == 1 and matching_starttime2:
                #     shift_timezone = matching_shift.timezone
                    
                #     # Convert start time based on shift timezone
                #     if shift_timezone == 'UTC+7':
                #         utc_plus_7_timezones = [tz for tz in all_timezones if timezone(tz).utcoffset(None) == utc_plus_7_offset]
                #         selected_timezone = timezone(random.choice(utc_plus_7_timezones))
                #     elif shift_timezone == 'GMT' or shift_timezone == 'GMT+8':
                #         gmt_or_gmt_plus_8_timezones = [tz for tz in all_timezones if tz.startswith('GMT') or timezone(tz).utcoffset(None) == gmt_plus_8_offset]
                #         selected_timezone = timezone(random.choice(gmt_or_gmt_plus_8_timezones))
                #     else:
                #         selected_timezone = timezone('UTC')

                #     # Assuming `matching_starttime` is a naive datetime object
                #     local_time = matching_starttime.replace(tzinfo=timezone('UTC')).astimezone(selected_timezone)
                #     local_time2 = matching_starttime2.replace(tzinfo=timezone('UTC')).astimezone(selected_timezone)
                #     ws.cell(row=current_row, column=4).value = local_time.strftime("%I:%M %p")
                #     ws.cell(row=current_row, column=7).value = local_time2.strftime("%I:%M %p")
                # else:
                #     ws.cell(row=current_row, column=4).value = ''
                #     ws.cell(row=current_row, column=7).value = ''
                                
                # ws.cell(row=current_row, column=8).value = "8.00" if attendance_log else "0.00"
                
                # time_off_label = matching_log._get_time_off_type_label() if matching_log else ''
                # ws.cell(row=current_row, column=10).value = time_off_label
                
                # latemins = matching_log.late_minutes
                # ws.cell(row=current_row, column=11).value = latemins  # Placeholder for LATE, can be adjusted

                # for col_num in range(1, 12):
                #     ws.cell(row=current_row, column=col_num).border = thin_border
                # current_row += 1

            # Fill in the summary headers
            for col_num, header in enumerate(summary_headers, 12):
                cell = ws.cell(row=start_row, column=col_num)
                cell.value = header
                cell.alignment = center_align
                cell.font = bold_font
                cell.border = thin_border
                
                # Apply light blue color fill to "Regular WD" and its corresponding value
                if header.startswith("Regular WD"):
                    cell.fill = light_blue_fill
            
            # Fill in the summary values
            for col_num, value in enumerate(summary_values, 12):
                cell = ws.cell(row=start_row + 1, column=col_num)
                cell.value = value
                cell.alignment = center_align
                cell.border = thin_border
                if col_num == 12:  # Assuming Regular WD is always the first in the summary
                    cell.fill = light_blue_fill

            last_row = start_row + len(date_range) + 2
            ws.merge_cells(start_row=start_row+2, start_column=12, end_row=last_row, end_column=35)

            for row in range(start_row, last_row + 1):
                for col_num in range(12, 36):
                    ws.cell(row=row, column=col_num).border = thin_border

            # Fill in the summary values
            # for col_num, value in enumerate(summary_values, 12):
            #     cell = ws.cell(row=start_row + 1, column=col_num)
            #     cell.value = value
            #     cell.alignment = center_align
            #     cell.border = thin_border
            
            # # Apply light blue color fill to "Regular WD" value cell
            # if col_num == 12:  # Assuming Regular WD is always the first in the summary
            #     cell.fill = light_blue_fill

            #     # Dynamically merge cells for the summary below the header
            #     last_row = start_row + len(date_range) + 2  # Calculate the last row to merge
            #     ws.merge_cells(start_row=start_row+2, start_column=12, end_row=last_row, end_column=35)

            #     # Ensure borders are applied correctly until column AH
            #     for row in range(start_row, last_row + 1):
            #         for col_num in range(12, 36):  # Columns L to AH
            #             ws.cell(row=row, column=col_num).border = thin_border
                    
        # Generate the date range
        date_range = generate_date_range(start_date, end_date)

        # Fill in the report for each example employee
        current_start_row = 1
        for employee in employee_names:
            fill_employee_block(current_start_row, date_range, employee, summary_values)
            current_start_row += len(date_range) + 3  # Adjust row start for the next employee block

         # Save Excel file to binary field
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        # Encode data to base64
        data_base64 = base64.b64encode(data)

        name_file = "PH REPORT NEW ATTENDANCE LOG REPORT.xlsx"

        # Save Excel file to attachment
        attachment = self.env['ir.attachment'].create({
            'name': name_file,
            'type': 'binary',
            'datas': data_base64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return attachment
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s/%s' % (attachment.id, urllib.parse.quote(attachment.name)),
            'target': 'self'
        }