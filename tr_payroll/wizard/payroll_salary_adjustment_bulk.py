from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import openpyxl
from io import BytesIO
import logging

_logger = logging.getLogger(__name__)

class PayrollSalaryAdjustmentBulkWizard(models.TransientModel):
    _name = 'payroll.salary.adjustment.bulk.wizard'
    _description = 'Payroll Import Excel Employee Wizard'

    file = fields.Binary(string='File', required=True)

    def upload_and_confirm(self):
        if not self.file:
            raise UserError(_("Please upload an Excel file."))

        file_content = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active

        headers = [cell.value.strip() if cell.value else '' for cell in ws[1]]
        expected_headers = [
            'Adjustment Date', 'Department', 'Category', 'Status', 'Salary Adjustment Detail/Employee',
            'Salary Adjustment Detail/Employee Basic Salary', 'Salary Adjustment Detail/Adjustment Type',
            'Salary Adjustment Detail/Adjustment Amount', 'Salary Adjustment Detail/Adjustment Reason',
        ]

        if headers != expected_headers:
            raise UserError(_("Invalid Excel file format. Please make sure the headers match the required format."))

        created_headers = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            adjustment_date = row[0]
            department_name = row[1].strip() if isinstance(row[1], str) else ''
            category_name = row[2].strip() if isinstance(row[2], str) else ''
            status = row[3]
            employee_name = row[4].strip() if isinstance(row[4], str) else ''
            basic_salary = row[5]
            adjustment_type_name = row[6].strip() if isinstance(row[6], str) else ''
            adjustment_amount = row[7]
            adjustment_reason = row[8]
            
            department = self.env['payroll.device.department'].search([('name', '=', department_name)], limit=1)
            if not department:
                raise UserError(_("Department '%s' not found.") % department_name)

            category = self.env['payroll.employee.categories'].search([('name', '=', category_name)], limit=1)
            if not category:
                raise UserError(_("Category '%s' not found.") % category_name)

            _logger.info("Searching header with: Date=%s, Department=%s, Category=%s", adjustment_date, department.id, category.id)
            header = self.env['payroll.salary.adjustment.header'].search([
                ('adjustment_date', '=', adjustment_date),
                ('department_id', '=', department.id),
                ('category_id', '=', category.id),
            ], limit=1)

            if not header:
                # Create a new header if it doesn't exist
                header = self.env['payroll.salary.adjustment.header'].create({
                    'adjustment_date': adjustment_date,
                    'department_id': department.id,
                    'category_id': category.id,
                    'status': 'draft'
                })
                _logger.info("Created new header: %s", header.id)
                created_headers.append(header) 
            else:
                # Check if the adjustment details already exist for this header
                existing_details = self.env['payroll.salary.adjustment.detail'].search([
                    ('header_id', '=', header.id),
                    ('employee_id.name', '=', employee_name),
                ])

                if existing_details:
                    _logger.info("Salary adjustment details for employee '%s' already exist for this header.", employee_name)
                    continue  # Skip

                _logger.info("Reused existing header: %s", header.id)

            employee = self.env['payroll.employees'].search([('name', '=', employee_name)], limit=1)
            if not employee:
                raise UserError(_("Employee '%s' not found.") % employee_name)

            adjustment_type = '1' if adjustment_type_name.lower() == 'salary increment' else '2'

            self.env['payroll.salary.adjustment.detail'].create({
                'header_id': header.id,
                'employee_id': employee.id,
                'employee_basic_salary': basic_salary,
                'adjustment_type': adjustment_type,
                'adjustment_amount': adjustment_amount,
                'adjustment_reason': adjustment_reason,
            })

        for header in created_headers:
            header.action_done()

        _logger.info("Finished processing the Excel file and marking headers as done.")
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }

