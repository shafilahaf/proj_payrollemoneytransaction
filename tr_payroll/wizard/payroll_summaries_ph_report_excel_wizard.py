import openpyxl.styles
from odoo import api, fields, models
from odoo.exceptions import UserError, ValidationError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl
import io
from io import BytesIO
import base64
from datetime import datetime, timedelta
import logging
import urllib.parse
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

_logger = logging.getLogger(__name__)
class PayrollSummariesPHReportExcelWizard(models.TransientModel):
    _name = "payroll.summaries.ph.report.excel.wizardd"
    _description = "Payroll Summaries PH Report Excel Wizard"

    department_id = fields.Many2one("payroll.device.department", string="Department", required=True)
    category_id = fields.Many2one("payroll.employee.categories", string="Category", required=True)
    employee_id = fields.Many2one("payroll.employees", string="Employee")
    month = fields.Selection([
        ('1', 'January'),
        ('2', 'February'),
        ('3', 'March'),
        ('4', 'April'),
        ('5', 'May'),
        ('6', 'June'),
        ('7', 'July'),
        ('8', 'August'),
        ('9', 'September'),
        ('10', 'October'),
        ('11', 'November'),
        ('12', 'December')
    ], string='Month', required=True)
    year = fields.Char(string='Year', required=True)
    month_text = fields.Char(string='Month Text', compute='_compute_month_text')
    periode = fields.Selection([
        ('1', 'Periode 1'),
        ('2', 'Periode 2'),
    ], string='Periode', default="1", required=True)

    @api.depends('month')
    def _compute_month_text(self):
        for rec in self:
            rec.month_text = dict(self._fields['month'].selection).get(rec.month)

    def payroll_summaries_ph_internal(self):
        return self.get_data_paysum_ph(forinternal=True)

    def get_data_paysum_ph(self, forinternal):
        """
        Generate Payroll Summaries PH Internal Report Excel
        """
        # Get payroll summaries
        domain = []
        if self.department_id:
            domain.append(('employee_id.department_id', '=', self.department_id.id))
        if self.category_id:
            domain.append(('employee_id.category_id', '=', self.category_id.id))
        if self.employee_id:
            domain.append(('employee_id', '=', self.employee_id.id))
        if self.month:
            domain.append(('month', '=', self.month))
        if self.year:
            domain.append(('year', '=', self.year))
        if self.periode:
            domain.append(('periode', '=', self.periode))

        payroll_summaries = self.env['payroll.payroll.summaries'].search(domain)

        # Create Excel file
        wb = Workbook()

        # Merge first row for title and date
        ws = wb.active
        ws.title = "PAYROLL SLIP SUMMARIES"
        ws.sheet_view.showGridLines = False  # Hide gridlines
        # ws.merge_cells('A1:AH1')
        ws['A1'] = f"PAYROLL SLIP SUMMARIES PERIODE {self.periode} {(self.month_text).upper()}-{self.year}"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
        ws['A2'] = f"Department: {self.department_id.name}"
        ws['A3'] = f"Category: {self.category_id.name}"
        ws['A4'] = f"Print Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        header = [
            'EMPLOYEE',
            'NIK',
            'BASIC RATE',
            'BI',
            'DAILY RATE',
            'HOURLY RATE',
            'MINUTE RATE',
            'FIX OT HOURS',
            'ALLOWANCE',
            'RWD',
            'DAY OFF',
            'ABSENCE',
            'SICK',
            'LATE',
            'LEAVE & PERMISSION',
            'ND',
            'SH',
            'LH',
            'REG OT',
            'SH OT',
            'LH OT',
            'OT RECEIPT',
            'REG OT AMT',
            'ND AMT',
            'SH AMT',
            'LH AMT',
            'SH OT AMT',
            'LH OT AMT',
            'ADJ ADD',
            'LATE MINS',
            'ABSENCE AMT',
            'ADJ DEDUCT',
            'SSS',
            'PHIL-HEALTH',
            'PAG-IBIG',
            'SSS LOANS',
            'CALAMITY LOAN',
            'HDMF LOAN',
            'ERROR',
            'ABS DED. IN ALLOWANCE',
            'LATE DED. IN ALLOWANCE',
            'OVER BREAK DED.',
            'OVER BREAK MINS',
            'ALLOWANCE',
            'TOTAL ADD',
            'TOTAL DEDUCTION',
            'NETT',
            '13TH'
        ]
        ws.append(header)

        # Set color for row A5
        for cell in ws['J5:V5'][0]:
            cell.fill = openpyxl.styles.PatternFill(start_color="ebcc34", end_color="ebcc34", fill_type="solid") #YELLOW
              

        for cell in ws['W5:AC5'][0]:
            cell.fill = openpyxl.styles.PatternFill(start_color="50c940", end_color="50c940", fill_type="solid") #GREEN
              

        for cell in ws['AD5:AM5'][0]:
            cell.fill = openpyxl.styles.PatternFill(start_color="ebcc34", end_color="ebcc34", fill_type="solid")
              

        for cell in ws['AN5:AO5'][0]:
            cell.fill = openpyxl.styles.PatternFill(start_color="50c940", end_color="50c940", fill_type="solid")
              

        for cell in ws['AP5:AQ5'][0]:
            cell.fill = openpyxl.styles.PatternFill(start_color="ebcc34", end_color="ebcc34", fill_type="solid")
              

        ws['AR5'].fill = openpyxl.styles.PatternFill(start_color="50c940", end_color="50c940", fill_type="solid")

        for cell in ws['AU5:AV5'][0]:
            cell.fill = openpyxl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF")  # Set font color to white

        # Define the border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply border to the entire row 5
        for cell in ws[5]:
            cell.border = thin_border

        # Freeze pane 
        ws.freeze_panes = 'A6'

        # Set column widths ( autofit )
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        for ps in payroll_summaries: 
            salid = self.env['payroll.salary.components'].search([
                ('is_basic_salary', '=', True)
            ])
            basicsal = self.env['payroll.payroll.summary.detail'].search([
                ('payroll_summary_id', '=', ps.id),
                ('type_id', '=', salid.id)
            ])
            overtimeid = self.env['payroll.salary.components'].search([
                ('ph_salary_type', '=', '1')
            ])
            overtimedata = self.env['payroll.payroll.summary.detail'].search([
                ('payroll_summary_id', '=', ps.id),
                ('type_id', '=', overtimeid.id)
            ])
            allowanceid = self.env['payroll.salary.components'].search([
                ('ph_salary_type', '=', '2')
            ])
            allowancedata = self.env['payroll.payroll.summary.detail'].search([
                ('payroll_summary_id', '=', ps.id),
                ('type_id', '=', allowanceid.id)
            ])
            
            attlogdata = self.env['payroll.attendance.log'].search([
                ('employee_id', '=', ps.employee_id.id),
                ('date', '>=', ps.from_date),
                ('date', '<=', ps.to_date)
            ])
            
            latemin = sum(att.late_minutes for att in attlogdata)
            
            attlognd = attlogdata.search([
                ('is_night_diff', '=', True),
                ('time_off_type', '=', False)
            ])
            ndhours = 0
            for attlog in attlognd:
                shift = self.env['payroll.shifts'].search([
                    ('id', '=', attlog.shift_id.id)
                ])
                
                if forinternal:
                    ndhours = ndhours + 8
                else:
                    shift_start_hours = int(shift.start_time)
                    shift_start_time = timedelta(hours=shift_start_hours)
                    if shift_start_time == 19:
                        ndhours = ndhours + 6
                    elif shift_start_time == 21:
                        ndhours = ndhours + 7
                    elif shift_start_time == 22:
                        ndhours = ndhours + 8
                
            SHHours_len = attlogdata.search_count([
                ('holiday_type', '=', '2')
            ]) * 8
            LHHours_len = attlogdata.search_count([
                ('holiday_type', '=', '1')
            ]) * 8
            RegOTDayHours_len = attlogdata.search_count([
                ('holiday_type', '=', False),
                ('time_off_type', '=', False)
            ]) * 3
            SHOTDayHours_len = attlogdata.search_count([
                ('holiday_type', '=', '2')
            ]) * 3
            LHOTDayHours_len = attlogdata.search_count([
                ('holiday_type', '=', '1')
            ]) * 3
            
            LHHours_len = attlogdata.search_count([
                ('holiday_type', '=', '1')
            ]) * 8
            
            dailyrate = 0
            ovetimeamount = 0
            allowanceamount = 0
            
            if overtimedata:
                ovetimeamount = overtimedata.amount / 26 / 3
                
            if allowancedata and self.periode == "1":
                allowanceamount = allowancedata.amount
            else:
                allowanceamount = 0
                
            if basicsal:
                dailyrate = basicsal.amount / 26.08
                hourlyrate = basicsal.amount / 26.08 / 8
                minuterate = basicsal.amount / 26.08 / 8 / 60

                mistake_entries = self.env['payroll.mistake.entries.details'].search([
                    ('employee_id', '=', ps.employee_id.id),
                    ('header_id.date', '>=', ps.from_date),
                    ('header_id.date', '<=', ps.to_date)
                ])

                deductions = self.env['payroll.deductions'].search([('ph_deduction_type', '!=', False)])

                deduct_amounts = [0] * 11

                for deduction in deductions:
                    deduct_type = int(deduction.ph_deduction_type)
                    deduct_amounts[deduct_type] = sum(
                        entry.amount for entry in mistake_entries if entry.deduction_id.ph_deduction_type == deduction.ph_deduction_type
                    )

                DeductAm1 = deduct_amounts[1]
                DeductAm2 = deduct_amounts[2]
                DeductAm3 = deduct_amounts[3]
                DeductAm4 = deduct_amounts[4]
                DeductAm5 = deduct_amounts[5]
                DeductAm6 = deduct_amounts[6]
                DeductAm7 = deduct_amounts[7]
                DeductAm8 = deduct_amounts[8]
                DeductAm9 = deduct_amounts[9]
                DeductAm10 = deduct_amounts[10]
                
                total_add = (
                    (basicsal.amount / 2) +
                    (allowanceamount - (ps.days_absent * 800) - (ps.late * 100)) +
                    ((RegOTDayHours_len + SHOTDayHours_len + LHOTDayHours_len) * ovetimeamount) +
                    ((hourlyrate / 10) * ndhours) +
                    (hourlyrate * SHHours_len * 0.3) +
                    (hourlyrate * LHHours_len) +
                    (hourlyrate * SHOTDayHours_len * 1.3 * 0.3) +
                    (hourlyrate * LHOTDayHours_len * 2 * 0.3) +
                    deduct_amounts[1]
                )
                
                totaldeduction = (
                    minuterate * latemin +
                    dailyrate * ps.days_absent +
                    DeductAm2 +
                    DeductAm3 +
                    DeductAm4 +
                    DeductAm5 +
                    DeductAm6 +
                    DeductAm7 +
                    DeductAm8 +
                    DeductAm9
                )
        

                row = [
                    ps.employee_id.name,
                    ps.employee_id.nik,
                    basicsal.amount,
                    (basicsal.amount / 2), # Rate BI
                    dailyrate, # Daily Rate
                    hourlyrate, # Hourly Rate
                    minuterate, # Minute Rate
                    ovetimeamount, # Fix OT Hours
                    allowanceamount, # Allowance
                    ps.total_point, #RWD / Reguler WD
                    ps.days_off, # Days Off
                    ps.days_absent, # Absensce 
                    ps.sick_permission, # Sick
                    latemin, #Late
                    (ps.days_leave + ps.days_permission), # Leave and Permission
                    ndhours, #ND
                    SHHours_len, # SHHours
                    LHHours_len,# LHHours
                    RegOTDayHours_len, # RegOTDayHours
                    SHOTDayHours_len, # SHOTDayHours
                    LHOTDayHours_len, # LHOTDayHours
                    (RegOTDayHours_len + SHOTDayHours_len + LHOTDayHours_len), # OT Receipt
                    
                    # CONTINUE CODE -- TRWahid
                    # CONTINUE CODE -- TRWahid
                    
                    (RegOTDayHours_len + SHOTDayHours_len + LHOTDayHours_len) * ovetimeamount, # REG OT AMT
                    (hourlyrate / 10) * ndhours, # ND AMT
                    hourlyrate * SHHours_len * 0.3, # SH AMT
                    hourlyrate * LHHours_len, # LH AMT
                    hourlyrate * SHOTDayHours_len * 1.3 * 0.3, # SH OT AMT
                    hourlyrate * LHOTDayHours_len * 2 * 0.3, # LH OT AMT
                    DeductAm1, # ADJ ADD
                    minuterate * latemin, # LATE MINS
                    dailyrate * ps.days_absent, # ABSENCE AMT
                    DeductAm2, # ADJ DEDUCT
                    DeductAm3, # SSS
                    DeductAm4, # PHIL-HEALTH
                    DeductAm5, # PAG-IBIG
                    DeductAm6, # SSS LOANS
                    DeductAm7, # CALAMITY LOAN
                    DeductAm8, # HDMF LOAN
                    DeductAm10, # ERROR
                    ps.days_absent * 800, # ABS DED. IN ALLOWANCE
                    ps.late, # LATE DED. IN ALLOWANCE
                    DeductAm9, # OVER BREAK DED.
                    0, # OVER BREAK MINS
                    allowanceamount - (ps.days_absent * 800) - (ps.late * 100), # ALLOWANCE
                    ((basicsal.amount / 2) +
                    (allowanceamount - (ps.days_absent * 800) - (ps.late * 100)) +
                    ((RegOTDayHours_len + SHOTDayHours_len + LHOTDayHours_len) * ovetimeamount) +
                    ((hourlyrate / 10) * ndhours) +
                    (hourlyrate * SHHours_len * 0.3) +
                    (hourlyrate * LHHours_len) +
                    (hourlyrate * SHOTDayHours_len * 1.3 * 0.3) +
                    (hourlyrate * LHOTDayHours_len * 2 * 0.3) +
                    deduct_amounts[1]), # TOTAL ADD
                    (minuterate * latemin +
                    dailyrate * ps.days_absent +
                    DeductAm2 +
                    DeductAm3 +
                    DeductAm4 +
                    DeductAm5 +
                    DeductAm6 +
                    DeductAm7 +
                    DeductAm8 +
                    DeductAm9), # TOTAL DEDUCTION
                    total_add - totaldeduction, # NETT
                    (basicsal.amount / 2) - (minuterate * latemin) - (dailyrate * ps.days_absent) # 13TH
                ]
            ws.append(row)

        # Save Excel file to binary field
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        # Encode data to base64
        data_base64 = base64.b64encode(data)

        name_file = f"PAYROLL SUMMARY PH INTERNAL PERIODE {self.periode} - {self.month_text}-{self.year}.xlsx"

        # Save Excel file to attachment
        attachment = self.env['ir.attachment'].create({
            'name': name_file,
            'type': 'binary',
            'datas': data_base64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return attachment
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s/%s' % (attachment.id, urllib.parse.quote(attachment.name)),
            'target': 'self'
        }
    
# Alfian Code
    def payroll_summaries_ph_external(self):
        """
        Generate Payroll Summaries PH External Report Excel with Headers Only
        """
        
        domain = []
        if self.department_id:
            domain.append(('employee_id.department_id', '=', self.department_id.id))
        if self.category_id:
            domain.append(('employee_id.category_id', '=', self.category_id.id))
        if self.employee_id:
            domain.append(('employee_id', '=', self.employee_id.id))
        if self.month:
            domain.append(('month', '=', self.month))
        if self.year:
            domain.append(('year', '=', self.year))
        if self.periode:
            domain.append(('periode', '=', self.periode))

        payroll_summaries = self.env['payroll.payroll.summaries'].search(domain)
        
        from_date = min(payroll_summaries.mapped('from_date')) if payroll_summaries else None
        to_date = max(payroll_summaries.mapped('to_date')) if payroll_summaries else None

        # Create Excel file
        wb = Workbook()

        # Merge first row for title and date
        ws = wb.active
        ws.title = "PAYROLL SLIP SUMMARIES"
        ws.sheet_view.showGridLines = False  # Hide gridlines
        ws.merge_cells('A1:AL1')
        ws['A1'] = "PAYROLL SLIP SUMMARIES"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Subtitle
        ws.merge_cells('A2:AL2')
        ws['A2'] = f"{self.month_text} - {self.year}"
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells('A3:AL3')
        if from_date and to_date:
            ws['A3'] = f"Periode {from_date.strftime('%d %b %Y')} - {to_date.strftime('%d %b %Y')}"
        else:
            ws['A3'] = "Periode Not Available"
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
        
        center_align = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            ('No.', 'Employee Name', 'BASIC RATE', 'Rate (BI MONTHLY)', '', 'Daily Rate', 'Hourly Rate', 'Regular WD', 'VL', 'SL', 'Late/UT', 'Day/s', 'VL Amount', 'SL Amount', 'Late/UT', 'Day/s Absent', 'Total Amount', 'OVERTIME HOURS', '', '', '', 'HOLIDAY OVERTIME HOURS', '', '', '', '', '', 'OVERTIME AMOUNT', '', '', '', 'HOLIDAY OVERTIME AMOUNT', '', '', '', '', '', 'OT TOTAL', 'HOLIDAY TOTAL', 'Adjusment(+/-)', 'GROSS', 'MANDATORY DEDUCTIONS', '', '',''),
            ('', '', '', '(BI MONTHLY)', '', '', '', '', '', '', 'Mins.', 'Absent', '', '', 'Amount', 'Amount', '', 'Reg. OT Hrs.', 'RD Hrs.', 'RDOT Hrs.', 'ND Hrs', 'SP Hol Hrs.', 'OT SP Hol Hrs.', 'RD/SP Hol Hrs.', 'Reg Hol Hrs.', 'OT Reg Hol Hrs.', 'RD Reg. Hol Hrs.', 'Reg OT amt.', 'RD amt.', 'RDOT amt.', 'ND amt.', 'Special Hol amt.', 'OT Special Hol amt.', 'RD/Special Hol amt.', 'Regular Hol amt.', 'OT/Regular Hol amt.', 'RD/Regular Hol amt.', 'OT Total', 'Holiday Total', 'Adjustment(+/-)', 'GROSS', 'SSS', 'PH HEALTH', 'PAG IBIG', 'WTAX')
        ]

        # Write headers
        for row_index, header_row in enumerate(headers, start=6):
            for col_index, header in enumerate(header_row, start=1):
                cell = ws.cell(row=row_index, column=col_index)
                cell.value = header
                cell.alignment = Alignment(horizontal='center', vertical='center')

        reg_ot_col = headers[1].index('Reg. OT Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=1.25)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('RD Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=1.3)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('RDOT Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=1.69)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('ND Hrs') + 1
        ws.cell(row=8, column=reg_ot_col, value=0.1)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('SP Hol Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=0.3)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('OT SP Hol Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=1.69)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('RD/SP Hol Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=1.5)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('Reg Hol Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=1)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('OT Reg Hol Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=2.6)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        reg_ot_col = headers[1].index('RD Reg. Hol Hrs.') + 1
        ws.cell(row=8, column=reg_ot_col, value=2.6)
        ws.cell(row=8, column=reg_ot_col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Define the border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Merge cells for the headers that span multiple rows
        ws.merge_cells('A6:A8')
        ws.merge_cells('B6:B8')
        ws.merge_cells('C6:C8')
        ws.merge_cells('D6:E8')  
        ws.merge_cells('F6:F8')
        ws.merge_cells('G6:G8')
        ws.merge_cells('H6:H8')
        ws.merge_cells('I6:I8')
        ws.merge_cells('J6:J8')
        ws.merge_cells('K6:K8')
        ws.merge_cells('L6:L8')
        ws.merge_cells('M6:M8')
        ws.merge_cells('N6:N8')
        ws.merge_cells('O6:O8')
        ws.merge_cells('P6:P8')
        ws.merge_cells('Q6:Q8')
        ws.merge_cells('R6:U6')  
        ws.merge_cells('V6:AA6')  
        ws.merge_cells('AB6:AE6') 
        ws.merge_cells('AF6:AK6')  
        ws.merge_cells('AL6:AL8')  
        ws.merge_cells('AM6:AM8') 
        ws.merge_cells('AN6:AN8')  
        ws.merge_cells('AO6:AO8') 
        ws.merge_cells('AP6:AS6') 

        # Define the range of rows to add data
        num_rows_to_add = 50  # Change as necessary

        max_row = ws.max_row
        for row in range(8, max_row + 1):
            ws.merge_cells(f'D{row}:E{row}')

        # Apply the border to the merged cells
        merged_ranges = ['A6:A8', 'B6:B8', 'C6:C8', 'D6:D7', 'E6:E8']
        for merged_range in merged_ranges:
            for row in ws[merged_range]:
                for cell in row:
                    cell.border = thin_border

        # Apply border to the entire header rows
        for row in ws.iter_rows(min_row=6, max_row=8):
            for cell in row:
                cell.border = thin_border

        # Apply borders to all cells from row 9 onward
        for row in range(9, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
            
        # Freeze pane 
        ws.freeze_panes = 'A10'

        # Set column widths ( autofit )
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width 
        
        current_row = 9  
        for idx, ps in enumerate(payroll_summaries, start=1):
            salid = self.env['payroll.salary.components'].search([
                ('is_basic_salary', '=', True)
            ])
            basicsal = self.env['payroll.payroll.summary.detail'].search([
                ('payroll_summary_id', '=', ps.id),
                ('type_id', '=', salid.id)
            ])
            overtimeid = self.env['payroll.salary.components'].search([
                ('ph_salary_type', '=', '1')
            ])
            overtimedata = self.env['payroll.payroll.summary.detail'].search([
                ('payroll_summary_id', '=', ps.id),
                ('type_id', '=', overtimeid.id)
            ])
            allowanceid = self.env['payroll.salary.components'].search([
                ('ph_salary_type', '=', '2')
            ])
            allowancedata = self.env['payroll.payroll.summary.detail'].search([
                ('payroll_summary_id', '=', ps.id),
                ('type_id', '=', allowanceid.id)
            ])
            
            attlogdata = self.env['payroll.attendance.log'].search([
                ('employee_id', '=', ps.employee_id.id),
                ('date', '>=', ps.from_date),
                ('date', '<=', ps.to_date)
            ])
            
            latemin = sum(att.late_minutes for att in attlogdata)
            
            attlognd = attlogdata.search([
                ('is_night_diff', '=', True),
                ('time_off_type', '=', False)
            ])
            ndhours = 0
            for attlog in attlognd:
                shift = self.env['payroll.shifts'].search([
                    ('id', '=', attlog.shift_id.id)
                ])
                shift_start_hours = int(shift.start_time)
                shift_start_time = timedelta(hours=shift_start_hours)
                if shift_start_time == 19:
                    ndhours = ndhours + 6
                elif shift_start_time == 21:
                    ndhours = ndhours + 7
                elif shift_start_time == 22:
                    ndhours = ndhours + 8
            
            dailyrate = 0
            ovetimeamount = 0
            allowanceamount = 0
            
            if overtimedata:
                ovetimeamount = overtimedata.amount / 26 / 3
                
            if allowancedata and self.periode == "1":
                allowanceamount = allowancedata.amount
            else:
                allowanceamount = 0
                
            if basicsal:
                dailyrate = basicsal.amount / 26.08
                hourlyrate = basicsal.amount / 26.08 / 8
                minuterate = basicsal.amount / 26.08 / 8 / 60

                mistake_entries = self.env['payroll.mistake.entries.details'].search([
                    ('employee_id', '=', ps.employee_id.id),
                    ('header_id.date', '>=', ps.from_date),
                    ('header_id.date', '<=', ps.to_date)
                ])

                deductions = self.env['payroll.deductions'].search([('ph_deduction_type', '!=', False)])

                deduct_amounts = [0] * 11

                for deduction in deductions:
                    deduct_type = int(deduction.ph_deduction_type)
                    deduct_amounts[deduct_type] = sum(
                        entry.amount for entry in mistake_entries if entry.deduction_id.ph_deduction_type == deduction.ph_deduction_type
                    )

                DeductAm1 = deduct_amounts[1]
                DeductAm2 = deduct_amounts[2]
                DeductAm3 = deduct_amounts[3]
                DeductAm4 = deduct_amounts[4]
                DeductAm5 = deduct_amounts[5]
                DeductAm6 = deduct_amounts[6]
                DeductAm7 = deduct_amounts[7]
                DeductAm8 = deduct_amounts[8]
                DeductAm9 = deduct_amounts[9]
                DeductAm10 = deduct_amounts[10]
                
                ndhours_len = (ndhours * dailyrate) * 0.1
                
                SHHours_len = attlogdata.search_count([
                    ('holiday_type', '=', '2')
                ]) * 8
                SHSundayHours_len = (attlogdata.search_count([
                    ('holiday_type', '=', 'LH'),
                    ('weekday', '=', '0')
                ]) * 8 * dailyrate) * 1.69
                LHHours_len = attlogdata.search_count([
                    ('holiday_type', '=', '1')
                ]) * 8
                LHSundayHours_len = (attlogdata.search_count([
                    ('holiday_type', '=', 'LH'),
                    ('weekday', '=', '0')
                ]) * 8 * dailyrate) * 1.5
                RegOTDayHours_len = attlogdata.search_count([
                    ('holiday_type', '=', False),
                    ('time_off_type', '=', False)
                ]) * 3
                SHOTDayHours_len = attlogdata.search_count([
                    ('holiday_type', '=', '2')
                ]) * 3
                
                LHOTDayHours_len = attlogdata.search_count([
                    ('holiday_type', '=', '1')
                ]) * 3
                
                LHHours_len = attlogdata.search_count([
                    ('holiday_type', '=', '1')
                ]) * 8
                
                total_add = (
                    (basicsal.amount / 2) +
                    (allowanceamount - (ps.days_absent * 800) - (ps.late * 100)) +
                    ((RegOTDayHours_len + SHOTDayHours_len + LHOTDayHours_len) * ovetimeamount) +
                    ((hourlyrate / 10) * ndhours) +
                    (hourlyrate * SHHours_len * 0.3) +
                    (hourlyrate * LHHours_len) +
                    (hourlyrate * SHOTDayHours_len * 1.3 * 0.3) +
                    (hourlyrate * LHOTDayHours_len * 2 * 0.3) +
                    deduct_amounts[1]
                )
                
                totaldeduction = (
                    minuterate * latemin +
                    dailyrate * ps.days_absent +
                    DeductAm2 +
                    DeductAm3 +
                    DeductAm4 +
                    DeductAm5 +
                    DeductAm6 +
                    DeductAm7 +
                    DeductAm8 +
                    DeductAm9
                )

                total_len = ((ps.total_point * dailyrate) + ((ps.days_leave + ps.days_permission) * dailyrate) + (ps.sick_permission * dailyrate) - (ps.days_absent * dailyrate) - latemin)

                row = [
                    idx,
                    ps.employee_id.name,
                    basicsal.amount, # Basic Rate
                    (basicsal.amount / 2), # Rate BI
                    dailyrate, # Daily Rate
                    hourlyrate, # Hourly Rate
                    ps.total_point, # Reguler WD
                    (ps.days_leave + ps.days_permission), # VLCount
                    ps.sick_permission, #SLCount
                    latemin, #Late Mins
                    ps.days_absent, # Days Absence
                    ((ps.days_leave + ps.days_permission) * dailyrate), #VL Amount
                    (ps.sick_permission * dailyrate), # SLAmount
                    self.fnGetFromPayDetail(ps.id, 1),# Late Amount
                    self.fnGetFromPayDetail(ps.id, 2), # Absence Amount
                    ((ps.total_point * dailyrate) + ((ps.days_leave + ps.days_permission) * dailyrate)
                     + (ps.sick_permission * dailyrate) - (ps.days_absent * dailyrate)
                     - latemin), # TotalAmount
                    '', #Reg. OT Hrs.
                    '', #RD Hrs.
                    '', #RDOT Hrs.
                    ndhours, #ND Hrs
                    SHHours_len, #SP Hol Hrs.
                    '', #OT SP Hol Hrs.
                    SHSundayHours_len, #RD/SP Hol Hrs.
                    LHHours_len, #Reg Hol Hrs.
                    '', #OT Reg Hol Hrs.
                    SHSundayHours_len, #RD Reg. Hol Hrs.
                    '', #Reg OT amt.
                    '', #RD amt.
                    '', #RDOT amt.
                    ndhours_len, #ND amt.
                    SHHours_len, #Special Hol amt.
                    '', #OT Special Hol amt.
                    SHSundayHours_len, #RD/Special Hol amt.
                    LHHours_len, #Regular Hol amt.
                    '', #OT/Regular Hol amt.
                    LHSundayHours_len, #RD/Regular Hol amt.
                    ndhours_len, #OT Total
                    SHHours_len + LHHours_len + SHSundayHours_len + LHSundayHours_len, #Holiday Total
                    '', #Adjustment(+/-)
                    total_len + ndhours_len + SHHours_len + LHHours_len + SHSundayHours_len + LHSundayHours_len, #GROSS
                    self.fnGetFromPayDetail(ps.id, 3), #SSS
                    self.fnGetFromPayDetail(ps.id, 4), #PH HEALTH
                    self.fnGetFromPayDetail(ps.id, 5), #PAG IBIG
                    self.fnGetFromPayDetail(ps.id, 6) #WTAX
                    
                    
                ]
            ws.append(row)
            current_row += 1

        # Save Excel file to binary field
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        # Encode data to base64
        data_base64 = base64.b64encode(data)

        name_file = f"PAYROLL SUMMARY PH EXTERNAL PERIODE {self.periode} - {self.month_text}-{self.year}.xlsx"

        # Save Excel file to attachment
        attachment = self.env['ir.attachment'].create({
            'name': name_file,
            'type': 'binary',
            'datas': data_base64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return attachment
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s/%s' % (attachment.id, urllib.parse.quote(attachment.name)),
            'target': 'self'
        }
        
    def fnGetFromPayDetail(self, payid, phdeductype):
        deductamount = 0
        
        deductdata = self.env['payroll.deductions'].search([
            ('ph_deduction_type', '=', phdeductype)
        ])
        paydetaildata = self.env['payroll.payroll.summary.detail'].search([
            ('payroll_summary_id', '=', payid),
            ('type_id', 'in', deductdata.ids)
        ])
        
        if paydetaildata:
            deductamount = paydetaildata.net_amount
            
        return deductamount
        
# Alfian Code