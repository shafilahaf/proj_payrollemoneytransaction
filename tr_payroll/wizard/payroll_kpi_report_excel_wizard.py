from odoo import api, fields, models
from odoo.exceptions import UserError, ValidationError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl
import io
import base64
from datetime import datetime, timedelta
import logging
import urllib.parse

_logger = logging.getLogger(__name__)
class PayrollKPIReportExcelWizard(models.TransientModel):
    _name = "payroll.kpi.report.excel.wizard"
    _description = "Payroll KPI Report Excel Wizard"

    year = fields.Char(string="Year", required=True)
    semester = fields.Selection([
        ('1', 'Semester 1'),
        ('2', 'Semester 2')
    ], string='Semester', required=True)

    def calculate_kpi(self):
        """
        Func in C#
        fncalcattlog DONE
        fncalperformancelog DONE
        """
        year = int(self.year)
        if self.semester == "1":
            startdate = datetime(year, 1, 1)
            enddate = datetime(year, 6, 30)
        else:
            startdate = datetime(year, 7, 1)
            enddate = datetime(year, 12, 31)

        kpis = self.env['payroll.key.performance.index'].search([
            ('date', '>=', enddate),
        ])
        for kpi in kpis:
            # kpi_details = self.env['payroll.key.performance.index.detail'].search([('kpi_ids', '=', kpi.id)])
            # kpi_details.unlink()
            kpi.unlink()
        
        # Process each active employee
        employee = self.env['payroll.employees'].search([
            ('working_status', '=', '1')
        ])
        for emp in employee:
            log_point = self._calculate_attendance_log(emp.id, startdate, False, enddate, emp.category_id.id)
            perf_point = self._calculate_performance_log(emp.id, startdate, enddate, False, emp.category_id.id)
            if log_point != 0 or perf_point != 0:
                kpi = self.env['payroll.key.performance.index'].create({
                    'date': enddate,
                    'year': year,
                    'semester': self.semester,
                    'employee_id': emp.id,
                })

                kpi_total_att_log_point = self._calculate_attendance_log(emp.id, startdate, kpi.id ,enddate, emp.category_id.id)
                kpi_total_performance_point = self._calculate_performance_log(emp.id, startdate, enddate, kpi.id,emp.category_id.id)
                kpi_total_kpi_point = kpi_total_att_log_point + kpi_total_performance_point

                kpi.write({
                    'total_att_log_point': kpi_total_att_log_point,
                    'total_performance_point': kpi_total_performance_point,
                    'kpi_point': kpi_total_kpi_point,
                })
        return {
            'type': 'ir.actions.act_window',
            'name': 'Payroll KPIs',
            'res_model': 'payroll.key.performance.index',
            'view_mode': 'tree,form',
            'domain': [('date', '=', enddate)],
            'target': 'current',
        }

    def _calculate_attendance_log(self, employee_id, startdate, kpi_id,enddate, category_id):
        """
        Func in C# :
        fnGetAttLogPoint DONE
        fncalctotalworking DONE
        """
        point = 0
        point_month = [0] * 7
        attlog_point = [0] * 20

        category = self.env['payroll.employee.categories'].browse(category_id)
        percentage = category.attendance_log_percent or 0

        start = startdate
        end = start + timedelta(days=31) - timedelta(days=1)
        total_month = 0

        for i in range(1, 7):
            if i > 1:
                start = start + timedelta(days=31)
                end = start + timedelta(days=31) - timedelta(days=1)

            point_month[i] = 100
            attendance_logs = self.env['payroll.attendance.log'].search([
                ('employee_id', '=', employee_id),
                ('date', '>=', start),
                ('date', '<=', end)
            ])

            if attendance_logs:
                total_month += 1

                # Placeholder for getting attendance log points
                attlog_point = self._get_att_log_points()

                total_work_days = self._calculate_total_working_days(employee_id, start, end)
                if len(attendance_logs) - total_work_days > 0:
                    point_month[i] += (len(attendance_logs) - total_work_days) * attlog_point[1]

                for log in attendance_logs:
                    if log.time_off_type == "1":  # leave
                        point_month[i] += attlog_point[8]
                    elif log.time_off_type == "3":  # permission
                        point_month[i] += attlog_point[2]
                    elif log.time_off_type == "2":  # sick
                        point_month[i] += attlog_point[3]

                    if log.is_latelog:
                        point_month[i] += attlog_point[4]
                    elif log.is_late_2:
                        point_month[i] += attlog_point[5]
                    elif log.is_late_3:
                        point_month[i] += attlog_point[9]
                    elif log.is_late_4:
                        point_month[i] += attlog_point[10]

                    if log.working_hours == 0 and log.time_off_type == False:  # forget check out
                        point_month[i] += attlog_point[7]

                if self.env['payroll.key.performance.index'].browse(kpi_id):
                    self.env['payroll.key.performance.index.detail'].create({
                        'kpi_header': kpi_id,
                        'type': 1,
                        'type_text': "Attendance Log",
                        'start_date': start,
                        'end_date': end,
                        'total_point': point_month[i],
                        'percentage': percentage,
                        'kpi_point': point_month[i] * percentage / 100,
                    })

                point += point_month[i] * percentage / 100

        if point != 0 and total_month != 0:
            point = point / total_month

        return point
    
    def _get_att_log_points(self):
        log_points = [0] * 20  # Adjust size based on the actual number of log points

        logpoint_data = self.env['payroll.attendance.log.point'].search([])

        for log in logpoint_data:
            if 0 <= int(log.source) < len(log_points):
                log_points[int(log.source)] = log.point

        return log_points
    
    def _calculate_total_working_days(self, employee_id, startdate, enddate):
        employee = self.env['payroll.employees'].browse(employee_id)

        # Default total days between start and end dates
        total_days = (enddate - startdate).days + 1

        # Adjust based on active and inactive dates
        if employee.active_date and employee.active_date > startdate.date():
            startdate = employee.active_date
            total_days = (enddate - startdate).days + 1

        if employee.inactive_date and employee.inactive_date <= enddate:
            enddate = employee.inactive_date
            total_days = (enddate - startdate).days

        if employee.active_date and employee.inactive_date:
            if employee.active_date > startdate and employee.inactive_date <= enddate:
                startdate = employee.active_date
                enddate = employee.inactive_date
                total_days = (enddate - startdate).days

        return total_days

    def _calculate_performance_log(self, employee_id, startdate, enddate, kpi_id, category_id):
        point = 0
        point_month = [0] * 100  # Assuming 100 as a placeholder

        category = self.env['payroll.employee.categories'].browse(category_id)
        percentage = category.performance_percent or 0

        total_survey = 0

        # Search for appraisals in the given date range
        appraisals = self.env['payroll.employee.appraisals'].search([
            ('employee_assessed', '=', employee_id),
            ('assessed_date', '>=', startdate.date()),
            ('assessed_date', '<=', enddate.date())
        ])

        if len(appraisals) > 1:
            for appraisal in appraisals:
                total_survey += 1
                appraisal_header = self.env['payroll.employee.appraisals'].search([('employee_assessed', '=', employee_id),('id', '=', kpi_id)])
                if kpi_id:
                    self.env['payroll.key.performance.index.detail'].create({
                        'kpi_header': kpi_id,
                        'type': 2,
                        'type_text': "Performance",
                        'start_date': appraisal.date,
                        'end_date': appraisal.date,
                        'total_point': appraisal_header.total_score,
                        'percentage': percentage,
                        'kpi_point': appraisal_header.total_score * percentage / 100,
                    })

                point += appraisal.score * percentage / 100

            if point != 0 and total_survey != 0:
                point = point / total_survey

        return point

    def payroll_kpi_report(self):
        """
        Generate Payroll KPI Report Excel
        """

        # Get payroll kpi
        domain = []
        if self.year:
            domain.append(('year', '=', self.year))
        if self.semester:
            domain.append(('semester', '=', self.semester))

        payroll_kpi = self.env['payroll.key.performance.index'].search(domain)

        # Create Excel file
        wb = Workbook()

        ws = wb.active
        ws.title = "KEY PERFORMANCE INDEX"
        ws['A1'] = f"KEY PERFORMANCE INDEX"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
        ws['A2'] = f"YEAR: {self.year}"
        ws['A3'] = f"SEMESTER: {self.semester if self.semester else ''}"
        ws['A4'] = f"Print Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        header = [
            'EMPLOYEE',
            'NIK',
            'DEPARTMENT',
            'WEBSITE',
            'CATEGORY',
            'POSITION',
            'YEAR',
            'SEMESTER',
            'TOTAL ATT. LOG. POINT',
            'TOTAL PERFORMANCE POINT',
            'KPI POINT',
        ]
        ws.append(header)
    

        # Set color for row A5
        for cell in ws['A5':'AH5'][0]:
            cell.fill = openpyxl.styles.PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Freeze pane from first row until fifth row
        ws.freeze_panes = 'A6'

        # Set column widths ( autofit )
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        for ps in payroll_kpi: 
            ws.append([
                ps.employee_id.name,
                ps.employee_id.nik,
                ps.employee_id.department_id.name,
                ps.employee_id.current_website.name,
                ps.employee_id.category_id.name,
                ps.employee_id.current_position.name,
                ps.year,
                ps.semester,
                f"{ps.total_att_log_point:.2f}" if ps.total_att_log_point is not None else '',
                f"{ps.total_performance_point:.2f}" if ps.total_performance_point is not None else '',
                f"{ps.kpi_point:.2f}" if ps.kpi_point is not None else '',
            ])

        # Save Excel file to binary field
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        # Encode data to base64
        data_base64 = base64.b64encode(data)

        name_file = ""
        if self.year or self.semester:
            name_file = f"KEY_PERFORMANCE_INDEX_{self.year}_SEMESTER {self.semester}.xlsx"

        # Save Excel file to attachment
        attachment = self.env['ir.attachment'].create({
            'name': name_file,
            'type': 'binary',
            'datas': data_base64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return attachment
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s/%s' % (attachment.id, urllib.parse.quote(attachment.name)),
            'target': 'self'
        }



