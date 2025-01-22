from odoo import api, fields, models
from odoo.exceptions import UserError, ValidationError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl
import io
import base64
from datetime import datetime, timedelta
import logging
import urllib.parse
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.styles import Alignment
from datetime import datetime


_logger = logging.getLogger(__name__)

class PayrollSummariesReportExcelWizard(models.TransientModel):
    _name = "payroll.summaries.report.excel.wizard"
    _description = "Payroll Summaries Report Excel Wizard"

    department_id = fields.Many2one("payroll.device.department", string="Department")
    category_id = fields.Many2one("payroll.employee.categories", string="Category")
    employee_id = fields.Many2one("payroll.employees", string="Employee")
    month = fields.Selection([
        ('01', 'January'),
        ('02', 'February'),
        ('03', 'March'),
        ('04', 'April'),
        ('05', 'May'),
        ('06', 'June'),
        ('07', 'July'),
        ('08', 'August'),
        ('09', 'September'),
        ('10', 'October'),
        ('11', 'November'),
        ('12', 'December')
    ], string='Month')
    # year = fields.Integer(string="Year", default=datetime.now().year)
    current_year = datetime.now().year
    year = fields.Selection(
        [(str(year), str(year)) for year in range(current_year - 1, current_year + 1)],  # 10 years before and after the current year
        string="Year"
    )
    month_text = fields.Char(string='Month Text', compute='_compute_month_text')

    @api.depends('month')
    def _compute_month_text(self):
        for rec in self:
            rec.month_text = dict(self._fields['month'].selection).get(rec.month)

    def payroll_summaries_report(self):
        """
        Generate Payroll Summaries Report Excel
        """

        # Get payroll summaries
        domain = []
        if self.department_id:
            domain.append(('employee_id.department_id', '=', self.department_id.id))
        if self.category_id:
            domain.append(('employee_id.category_id', '=', self.category_id.id))
        if self.employee_id:
            domain.append(('employee_id', '=', self.employee_id.id))
        if self.month:
            domain.append(('month', '=', self.month))
        if self.year:
            domain.append(('year', '=', self.year))

        payroll_summaries = self.env['payroll.payroll.summaries'].search(domain)

        # Create Excel file
        wb = Workbook()

        # Merge first row for title and date
        ws = wb.active
        ws.title = "Payroll Summaries Report"
        ws.sheet_view.showGridLines = False  # Hide gridlines
        ws['A1'] = f"Payroll Summaries Report {self.month_text} - {self.year}"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
        ws['A2'] = f"Department: {self.department_id.name}"
        ws['A3'] = f"Category: {self.category_id.name}"
        # ws['A4'] = f"Print Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        future_date = datetime.now() + timedelta(hours=7)
        ws['A4'] = f"Print Date: {future_date.strftime('%Y-%m-%d %H:%M:%S')}"

        # Define the header columns
        header = [
            'NIK',
            'EMPLOYEE',
            'DEPARTMENT',
            'WEBSITE - CATEGORY',
            'POSITION',
            'MONTH',
            'YEAR',
            'TOTAL WORKING DAYS',
            'DAYS IN',
            'DAYS ABSENCE',
            'DAYS PERMISSION',
            'DAYS LEAVES',
            'SICK PERMISSION',
            'DAY OFF',
            'LATE',
            'LATE 2',
            'FORGET CHECKOUT COUNT',
            'MISTAKE 1',
            'MISTAKE 2',
        ]
        
        # Create a set to keep track of unique NameDetails
        name_details = set()
        basic_salary_currencies = set()
        for ps in payroll_summaries:
            for detail in ps.payroll_summary_detail_ids:
                detail_str = f"{detail.name} - {detail.currency_id.name if detail.currency_id else ''}"
                name_details.add(detail_str)

                if detail.name == "Basic Salary":
                    currency = detail.currency_id.name if detail.currency_id else ''
                    basic_salary_currencies.add(currency)

        net_salary_currencies = set()
        for ps in payroll_summaries:
            for detail in ps.payroll_summary_detail_ids:
                currency = detail.currency_id.name if detail.currency_id else ''
                net_salary_currencies.add(currency)

        # Sort currencies for consistent column ordering
        basic_salary_currencies = sorted(basic_salary_currencies)
        net_salary_currencies = sorted(net_salary_currencies)

        original_basic_salary_amounts = self.env['payroll.employee.salary'].search([
            ('is_basic_salary', '=', True),
        ])

        og_set = set()
        for obsa in original_basic_salary_amounts:
            currency = obsa.currency_id.name if obsa.currency_id else ''
            og_set.add(f"Original Basic Salary Amount - {currency}")

        # Add Original Basic Salary Amounts to the header
        for obsa in og_set:
            header.append(obsa)
            
 
        def custom_sort_key(detail_str, ps):
            for detail in ps.payroll_summary_detail_ids:
                if detail_str.startswith(detail.name):  # Match the detail by its name
                    if "Original" in detail_str:
                        return ('A', detail_str)  # 'A' will be at the top
                    elif detail.type == 1:
                        return ('B', detail_str)
                    elif detail.type == 2:
                        return ('C', detail_str)
                    elif detail.type == 3:
                        return ('D', detail_str)
                    else:
                        return ('Z', detail_str)  # Default for other types
            
            return ('Z', detail_str)  # Default fallback

        # name_details = sorted(name_details) #, key=sort_by_class
        name_details = sorted(name_details, key=lambda detail_str: custom_sort_key(detail_str, payroll_summaries))
        header.extend(name_details)
        for currency in net_salary_currencies:
            header.append(f"Net Salary - {currency}")
        header.append('BANK')
        header.append('BANK ACCOUNT NO.')
        ws.append(header)

        # Define Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set color for row A5
        last_column_letter = get_column_letter(len(header))
        for cell in ws[f"A5:{last_column_letter}5"][0]:
            cell.fill = openpyxl.styles.PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.border = thin_border

        # Freeze pane 
        ws.freeze_panes = 'A6'

        ws.column_dimensions['A'].width = 10  # Width for NIK column
        ws.column_dimensions['B'].width = 40  # Width for Name column

        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            if column not in ['A', 'B']:  
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        for ps in payroll_summaries:
            row = [
                ps.employee_id.nik,
                ps.employee_id.name,
                ps.employee_id.department_id.name,
                ps.employee_id.current_website.name + ' - ' + ps.employee_id.category_id.name if ps.employee_id.current_website else ps.employee_id.category_id.name,
                ps.employee_id.current_position.name,
                ps.month,
                ps.year,
                ps.total_work_days,
                ps.days_in,
                ps.days_absent,
                ps.days_permission,
                ps.days_leave,
                ps.sick_permission,
                ps.days_off,
                ps.late,
                ps.late2,
                ps.forget_checkout_count,
                ps.mistake_1_count,
                ps.mistake_2_count,
            ]
            
            thousand_separator_format = "#,##0.00"
            for row_index in range(6, ws.max_row + 1):
                for col_index in range(len(header)):
                    cell = ws.cell(row=row_index, column=col_index + 1)
                    if isinstance(cell.value, (int, float)):
                        if col_index in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]: 
                            continue
                        cell.number_format = thousand_separator_format


            # Original 
            original_basic_salary_amounts_dict = {f"Original Basic Salary Amount - {obsa.currency_id.name if obsa.currency_id else ''}": obsa.amount
                                                for obsa in original_basic_salary_amounts if obsa.employee_id == ps.employee_id}
            
            for obsa_header in og_set:
                value = original_basic_salary_amounts_dict.get(obsa_header, None)
                if not value:
                    value = 0
                row.append(value)
                # formatted = f"{value:,.2f}"
                # row.append(formatted)
            # Original 

            # Salary Details
            name_details_dict = {f"{detail.name} - {detail.currency_id.name if detail.currency_id else ''}": detail.net_amount
                                for detail in ps.payroll_summary_detail_ids}

            for detail_header in name_details:
                value = name_details_dict.get(detail_header, None)
                if not value:
                    value = 0
                row.append(value)
                # formatted = f"{value:,.2f}"
                # row.append(formatted)
            # Salary Details

            # Net Salary
            net_salary_per_currency = {currency: 0 for currency in net_salary_currencies}
            for detail in ps.payroll_summary_detail_ids:
                currency = detail.currency_id.name if detail.currency_id else ''
                net_salary_per_currency[currency] += detail.net_amount

            for currency in net_salary_currencies:
                row.append(net_salary_per_currency[currency])
                # formatted = f"{net_salary_per_currency[currency]:,.2f}"
                # row.append(formatted)
            # Net Salary

            row.append(ps.employee_id.bank_id.name if ps.employee_id.bank_id else '')
            row.append(ps.employee_id.bank_account_number if ps.employee_id.bank_account_number else '')

            ws.append(row)

            # Apply thin border to all cells in the current row
            last_column_letter = get_column_letter(len(row))
            for cell in ws[f"A{ws.max_row}:{last_column_letter}{ws.max_row}"][0]:
                cell.border = thin_border

        # Save Excel file to binary field
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        # Encode data to base64
        data_base64 = base64.b64encode(data)

        # Start with the basic part of the filename
        name_file = f"Payroll_Summary_{self.year}_{self.month_text}"

        if self.category_id:
            name_file += f"_{self.category_id.name}"

        if self.department_id:
            name_file += f"_{self.department_id.name}"

        if self.employee_id:
            name_file += f"_{self.employee_id.name}"

        name_file += ".xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': name_file,
            'type': 'binary',
            'datas': data_base64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return attachment
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s/%s' % (attachment.id, urllib.parse.quote(attachment.name)),
            'target': 'self'
        }