from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import openpyxl
from io import BytesIO

class PayrollImportExcelEmployeeWizard(models.TransientModel):
    _name = 'payroll.import.excel.employee.wizard'
    _description = 'Payroll Import Excel Employee Wizard'

    file = fields.Binary(string='File', required=True)

    def upload(self):
        """
        upload"""
        if not self.file:
            raise UserError(_("Please upload an Excel file."))

        file_content = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active

        # Check the headers in the first row
        headers = [cell.value.strip() for cell in ws[1]]
        expected_headers = [
            'Name', 'NIK', 'EmployeeStatusId', 'CompanyId', 'CategoryId', 'CurrentPosition',
            'Department', 'ContractId', 'ContractStartDate', 'VisaNo', 'VisaExpireDate',
            'ActiveDate', 'BankAccountId', 'BankAccountNo', 'Address', 'Address2', 'PhoneNo',
            'Email', 'EmergencyContact', 'EmergencyPhoneNo', 'GenderText', 'DateofBirth',
            'CountryofBirthId', 'PlaceofBirthId', 'PassportNo', 'ManagerId', 'Current Website'
        ]

        if headers != expected_headers:
            raise UserError(_("Invalid Excel file format. Please make sure the headers match the required format."))

        def get_id_or_raise(model_name, value):
            if not value:
                return False
            record = self.env[model_name].search([('name', '=', value)], limit=1)
            if not record:
                raise UserError(_(f"Record for {model_name} with name '{value}' not found."))
            return record.id

        for row in ws.iter_rows(min_row=2, values_only=True):
            emp_data = {
                'name': row[0],
                'nik': row[1],
                'employee_status_id': get_id_or_raise('payroll.employee.status', row[2]),
                'company_id': get_id_or_raise('payroll.companies', row[3]),
                'category_id': get_id_or_raise('payroll.employee.categories', row[4]),
                'current_position': get_id_or_raise('payroll.positions', row[5]),
                'department_id': get_id_or_raise('payroll.device.department', row[6]),
                'contract_id': get_id_or_raise('payroll.contracts', row[7]),
                'contract_start_date': row[8],
                'visa_number': row[9],
                'visa_expire_date': row[10],
                'active_date': row[11],
                'bank_id': get_id_or_raise('payroll.banks', row[12]),
                'bank_account_number': row[13],
                'address': row[14],
                'address_2': row[15],
                'phone': row[16],
                'email': row[17],
                'emergency_contact': row[18],
                'emergency_phone': row[19],
                'gender': '1' if row[20].lower() == 'male' else '2',
                'date_of_birth': row[21],
                'country_id': get_id_or_raise('payroll.countries', row[22]),
                'city_id': get_id_or_raise('payroll.cities', row[23]),
                'passport_number': row[24],
                'manager_id': get_id_or_raise('payroll.employees', row[25]),
                'working_status': '1',
                'current_website': get_id_or_raise('payroll.websites', row[26]),
            }

            # Create or update the employee record based on NIK
            employee = self.env['payroll.employees'].search([('nik', '=', emp_data['nik'])], limit=1)
            if employee:
                employee.write(emp_data)
            else:
                self.env['payroll.employees'].create(emp_data)

        # Optionally, you may want to add a success message or any other post-processing steps here.
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }